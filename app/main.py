from flask import Flask, render_template, request, redirect, url_for, flash, make_response
import io
import os
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import Workbook

app = Flask(__name__)

UPLOAD_FOLDER = 'app/data/uploads'
ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = 'ismael123456789'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/", methods=['GET', 'POST'])
def index():
    return render_template("index.html")

@app.route("/upload_dataset", methods=['POST'])
def upload_dataset():
  #route for the uploaded data
  if 'dataset' not in request.files:
    return redirect(request.url)
   
  file = request.files['dataset']
  if file.filename == '':
    return redirect(request.url)

  if file and allowed_file(file.filename):
    filename = secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    try:
        #Data processing based on file extension
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(file_path)
        else:
            #Unsupported file format
            return render_template('error.html', message="Unsupported file format. Only CSV, XLSX and XLS files are allowed.")
            
          #Processing complete, rendering the index.html
          #Passing DataFrame as JSON for potential usage in the template
        return render_template('index.html', message="Dataset uploaded and processed successfuly!", data=df.to_json())        
    except Exception as e:
            #Handle any exceptions during data processing
            return render_template('error.html', message=f"An error occured while processing the data: {str(e)}")
    else:
        return redirect(request.url)
    
@app.route("/view_data/<filename>", methods=['GET'])
def view_data(filename):
    # Construct the file path
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    # Check if the file exists
    if not os.path.exists(file_path):
        return render_template("error.html", message="File not found")

    try:
        # Read the uploaded dataset file
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path)
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            df = pd.read_excel(file_path)
        else:
            # Unsupported file format
            return render_template("error.html", message="Unsupported file format. Only CSV, XLSX, and XLS files are allowed.")
        
        # Pagination logic
        page = request.args.get('page', 1, type=int)
        per_page = 10  # Number of rows per page
        offset = (page - 1) * per_page
        data = df.iloc[offset:offset + per_page]
        
        # Pass the data and pagination information to the view_data.html template
        return render_template("view_data.html", data=data.to_html(index=False), page=page)
    except Exception as e:
        # Handle any exceptions
        return render_template("error.html", message=f"An error occurred: {str(e)}")
    
@app.route("/list_files", methods=['GET'])
def list_files():
    # Get the list of files in the UPLOAD_FOLDER directory
    files = os.listdir(app.config['UPLOAD_FOLDER'])
    num_files = len(files)
    
    return render_template("file_list.html", num_files=num_files, files=files)

@app.route("/export_data", methods=['POST'])
def export_data():
    # Check if processed data exists from upload_dataset
    if 'df' not in globals():
        return render_template("error.html", message="No data uploaded or processed yet.")

    export_format = request.form.get('export_format')
    if not export_format or export_format not in ('csv', 'xlsx'):
        return render_template("error.html", message="Invalid export format selected.")

    try:
        data_to_export = None
        filename = f"processed_data.{export_format}"

        # Export data based on chosen format
        if export_format == 'csv':
            data_to_export = globals()['df'].to_csv(index=False)
            content_type = 'text/csv; charset=utf-8'
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Processed Data"
            for col_idx, col_name in enumerate(df.columns):
                ws.cell(row=1, column=col_idx+1).value = col_name
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    ws.cell(row=row_idx+2, column=col_idx+1).value = value
            data_to_export = io.BytesIO()
            wb.save(data_to_export)
            data_to_export.seek(0)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        # Set response headers for the chosen format
        response = make_response(data_to_export)
        response.headers['Content-Type'] = content_type
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        return response

    except Exception as e:
        return render_template("error.html", message=f"An error occurred during export: {str(e)}")

@app.route("/delete_file/<filename>", methods=['POST'])
def delete_file(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(file_path):
        os.remove(file_path)
        flash("File deleted successfully", "success")
        # Redirect to file list after successful deletion
        return redirect(url_for('list_files'))
    else:
        flash("File not found", "error")
        # Redirect to file list even if file not found
        return redirect(url_for('list_files'))



if __name__ == "__main__":
    app.run(debug=True)